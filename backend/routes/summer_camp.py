"""
Summer Camp 2026 — Booking & Payment Routes
"""
import os
import uuid
import time
import logging
import asyncio
import io
from datetime import datetime, timezone
from typing import Optional, List
from fastapi import APIRouter, HTTPException, Depends, Request, UploadFile, File, BackgroundTasks
from fastapi.responses import FileResponse, StreamingResponse
from pydantic import BaseModel

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False

try:
    from cashfree_pg.models.create_order_request import CreateOrderRequest
    from cashfree_pg.api_client import Cashfree
    from cashfree_pg.models.customer_details import CustomerDetails as CashfreeCustomerDetails
    from cashfree_pg.models.order_meta import OrderMeta
    CASHFREE_AVAILABLE = True
except ImportError:
    CASHFREE_AVAILABLE = False
    Cashfree = None
    CashfreeCustomerDetails = None

from .shared import db, get_current_user

router = APIRouter()

CASHFREE_APP_ID = os.getenv("CASHFREE_APP_ID", "")
CASHFREE_SECRET_KEY = os.getenv("CASHFREE_SECRET_KEY", "")
CASHFREE_ENVIRONMENT = os.getenv("CASHFREE_ENVIRONMENT", "SANDBOX")
CASHFREE_API_VERSION = "2023-08-01"
CAMP_PRICE = 1999.0

# ── WhatsApp Enrollment Helper ────────────────────────────────────────────────
async def _build_enrolled_wa_params(booking: dict) -> list:
    """Build correct params list for the summercamp_enrolled WhatsApp template.
    Template expects: {1}=child_name, {2}=batch_dates, {3}=center_address, {4}=wa_group_link
    """
    first_name = (booking.get("child_name") or "").split()[0] if booking.get("child_name") else "there"

    # Batch dates: use stored batch_dates field, or fall back to BATCH_DATES lookup
    batch_dates = booking.get("batch_dates") or ""
    if not batch_dates:
        bd = BATCH_DATES.get(booking.get("batch_week", ""), {})
        batch_dates = bd.get("weekday") or booking.get("batch_week", "")

    # Center address: look up in centers collection by center slug/id
    center_address = booking.get("center_label") or booking.get("center") or ""
    wa_group_link = ""
    try:
        cval = booking.get("center", "")
        center_rec = await db.centers.find_one(
            {"$or": [{"id": cval}, {"slug": cval}, {"name": {"$regex": cval, "$options": "i"}}]},
            {"_id": 0, "address": 1, "wa_group_link": 1}
        )
        if center_rec:
            center_address = center_rec.get("address") or center_address
            wa_group_link = center_rec.get("wa_group_link") or ""
    except Exception:
        pass

    if not wa_group_link:
        wa_group_link = os.environ.get("SUMMERCAMP_WA_GROUP_LINK", "Contact your center for the WhatsApp group link")

    return [first_name, batch_dates or "See confirmation email for dates", center_address, wa_group_link]

if CASHFREE_AVAILABLE and CASHFREE_APP_ID and CASHFREE_SECRET_KEY:
    Cashfree.XClientId = CASHFREE_APP_ID
    Cashfree.XClientSecret = CASHFREE_SECRET_KEY
    if CASHFREE_ENVIRONMENT == "PRODUCTION":
        Cashfree.XEnvironment = Cashfree.PRODUCTION
    else:
        Cashfree.XEnvironment = Cashfree.SANDBOX

def get_cf_client():
    if not CASHFREE_AVAILABLE or not CASHFREE_APP_ID or not CASHFREE_SECRET_KEY:
        raise HTTPException(status_code=500, detail="Payment gateway not configured")
    cf_env = Cashfree.PRODUCTION if CASHFREE_ENVIRONMENT == "PRODUCTION" else Cashfree.SANDBOX
    return Cashfree(cf_env)

AGE_GROUPS = {
    "explorers": {"label": "Little Explorers", "ages": "4-8"},
    "creators": {"label": "Tech Creators", "ages": "9-12"},
    "innovators": {"label": "Future Innovators", "ages": "13-16"},
}

CENTERS = {
    "mira_road": "Mira Road",
    "dombivli": "Dombivli",
    "andheri": "Andheri West",
    "online": "Online",
}

async def _resolve_center_label(center_id: str) -> str:
    """Resolve a center ID (UUID or legacy slug) to a human-readable label (just the name)."""
    if not center_id:
        return ""
    doc = await db.centers.find_one({"id": center_id}, {"_id": 0, "name": 1})
    if doc:
        return doc.get("name", center_id)
    return CENTERS.get(center_id, center_id)

BATCH_DATES = {
    "week1": {"weekday": "May 4-8, 2026 (Mon-Fri)"},
    "week2": {"weekday": "May 11-15, 2026 (Mon-Fri)"},
    "week3": {"weekday": "May 18-22, 2026 (Mon-Fri)"},
    "week4": {"weekday": "May 25-29, 2026 (Mon-Fri)"},
}


class SummerCampRegistration(BaseModel):
    child_name: str
    parent_name: Optional[str] = ""
    parent_phone: str
    parent_email: str
    age_group: str  # explorers | creators | innovators
    batch_type: str = "weekday"  # weekday only
    batch_week: str  # week1 | week2 | week3 | week4
    mode: str       # offline | online
    center: str     # mira_road | dombivli | andheri | online
    payment_mode: str  # cashfree | cash


class BookingUpdate(BaseModel):
    child_name: Optional[str] = None
    parent_name: Optional[str] = None
    parent_phone: Optional[str] = None
    parent_email: Optional[str] = None


class StatusUpdate(BaseModel):
    crm_status: str  # phone_captured | lead | hot_lead | converted | payment_offline | lost_lead
    lost_reason: Optional[str] = None  # phone_not_picking | not_available_dates | location_too_far | other


class FollowupStatusUpdate(BaseModel):
    followup_status: str  # not_contacted | call_not_picked | call_cut | callback_requested | interested | non_serviceable
    callback_date: Optional[str] = None   # ISO date string when followup_status = callback_requested
    callback_time: Optional[str] = None


class CommentAdd(BaseModel):
    text: str
    author: Optional[str] = "Admin"
    comment_type: Optional[str] = "comment"  # "comment" | "call_done"
    call_date: Optional[str] = None   # ISO date when call was made
    call_time: Optional[str] = None   # HH:MM time of call


class PartialLeadCapture(BaseModel):
    parent_phone: str
    age_group: str
    batch_type: str
    batch_week: str
    mode: str
    center: str
    ref: Optional[str] = None          # tracking link slug


class CompleteLead(BaseModel):
    child_name: str
    parent_email: str
    payment_mode: str
    parent_name: Optional[str] = ""


class PaymentInitRequest(BaseModel):
    booking_id: str
    frontend_url: str = "https://oll.co"


class CreateTrackingLinkRequest(BaseModel):
    name: str


# ── helpers ────────────────────────────────────────────────────────────────────
import re

def _slugify(name: str) -> str:
    """Convert a name to a URL-safe slug."""
    slug = re.sub(r'[^a-z0-9]+', '-', name.lower().strip())
    return slug.strip('-')[:40]


async def _increment_tracking(slug: str, field: str):
    """Safely increment a tracking counter. Silently ignores missing slugs."""
    if not slug:
        return
    await db.summer_camp_tracking_links.update_one(
        {"slug": slug},
        {"$inc": {field: 1}}
    )


# ─────────────────────────── TRACKING LINK ENDPOINTS ─────────────────────────

@router.get("/summer-camp/tracking-links")
async def list_tracking_links(user: dict = Depends(get_current_user)):
    links = await db.summer_camp_tracking_links.find({}, {"_id": 0}).sort("created_at", -1).to_list(200)
    return links


@router.post("/summer-camp/tracking-links")
async def create_tracking_link(data: CreateTrackingLinkRequest, user: dict = Depends(get_current_user)):
    base_slug = _slugify(data.name)
    if not base_slug:
        raise HTTPException(status_code=400, detail="Name is required")
    # Ensure uniqueness
    slug = base_slug
    suffix = 1
    while await db.summer_camp_tracking_links.find_one({"slug": slug}):
        slug = f"{base_slug}-{suffix}"
        suffix += 1

    link_id = str(uuid.uuid4())
    doc = {
        "id": link_id,
        "name": data.name.strip(),
        "slug": slug,
        "views": 0,
        "leads": 0,
        "conversions": 0,
        "created_at": datetime.now(timezone.utc).isoformat(),
    }
    await db.summer_camp_tracking_links.insert_one(doc)
    return {k: v for k, v in doc.items() if k != '_id'}


@router.delete("/summer-camp/tracking-links/{link_id}")
async def delete_tracking_link(link_id: str, user: dict = Depends(get_current_user)):
    result = await db.summer_camp_tracking_links.delete_one({"id": link_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Link not found")
    return {"message": "Deleted"}


@router.post("/summer-camp/track-view/{slug}")
async def track_view(slug: str):
    """Called by the landing page when loaded via a tracking link."""
    await _increment_tracking(slug, "views")
    return {"ok": True}


# ───────────────────────────────────────────────────────────────────────────────


@router.post("/summer-camp/capture-lead")
async def capture_lead(data: PartialLeadCapture):
    """Save a partial lead at the phone number step — before full details are entered."""
    ref_num = str(int(time.time()))[-6:]
    booking_ref = f"SC2026-{ref_num}"
    booking_id = str(uuid.uuid4())

    batch_dates = BATCH_DATES.get(data.batch_week, {}).get(data.batch_type, "")

    # Resolve tracking link name for display
    source_name = "Direct"
    if data.ref:
        link = await db.summer_camp_tracking_links.find_one({"slug": data.ref}, {"_id": 0})
        if link:
            source_name = link.get("name", data.ref)
            await _increment_tracking(data.ref, "leads")

    center_label = await _resolve_center_label(data.center)

    doc = {
        "id": booking_id,
        "booking_ref": booking_ref,
        "parent_phone": data.parent_phone,
        "child_name": "",
        "parent_name": "",
        "parent_email": "",
        "age_group": data.age_group,
        "age_group_label": AGE_GROUPS.get(data.age_group, {}).get("label", ""),
        "age_group_ages": AGE_GROUPS.get(data.age_group, {}).get("ages", ""),
        "batch_type": data.batch_type,
        "batch_week": data.batch_week,
        "batch_dates": batch_dates,
        "mode": data.mode,
        "center": data.center,
        "center_label": center_label,
        "payment_mode": "cashfree",
        "amount": CAMP_PRICE,
        "payment_status": "pending",
        "crm_status": "phone_captured",
        "source_ref": data.ref or "",
        "source_name": source_name,
        "created_at": datetime.now(timezone.utc).isoformat(),
        "updated_at": datetime.now(timezone.utc).isoformat(),
    }
    await db.summer_camp_bookings.insert_one(doc)
    logging.info(f"Summer camp phone lead captured: {booking_ref} - {data.parent_phone} - source: {source_name}")
    return {"booking_id": booking_id, "booking_ref": booking_ref}


@router.patch("/summer-camp/complete-lead/{booking_id}")
async def complete_lead(booking_id: str, data: CompleteLead):
    """Update a phone-captured lead with full registration details."""
    booking = await db.summer_camp_bookings.find_one({"id": booking_id}, {"_id": 0})
    if not booking:
        raise HTTPException(status_code=404, detail="Booking not found")

    # ── Deduplication: same phone + same child name ───────────────────────────
    # Normalize for comparison (case-insensitive, trimmed)
    norm_child = (data.child_name or "").strip().lower()
    norm_phone = (booking.get("parent_phone") or "").strip()

    if norm_child and norm_phone:
        existing = await db.summer_camp_bookings.find_one(
            {
                "id": {"$ne": booking_id},
                "parent_phone": norm_phone,
                "child_name": {"$regex": f"^{norm_child}$", "$options": "i"},
                "crm_status": {"$in": ["lead", "hot_lead", "converted", "payment_offline"]},
            },
            {"_id": 0}
        )
        if existing:
            # Same person filled the form again — increment return_count, drop the duplicate
            return_count = existing.get("return_count", 1) + 1
            await db.summer_camp_bookings.update_one(
                {"id": existing["id"]},
                {"$set": {
                    "return_count": return_count,
                    "last_returned_at": datetime.now(timezone.utc).isoformat(),
                    "updated_at": datetime.now(timezone.utc).isoformat(),
                }}
            )
            # Remove the duplicate phone_captured entry
            await db.summer_camp_bookings.delete_one({"id": booking_id})
            logging.info(f"[SC] Duplicate lead suppressed for phone {norm_phone} / {data.child_name}. Existing: {existing['id']}, return_count={return_count}")
            return {
                "booking_id": existing["id"],
                "booking_ref": existing.get("booking_ref"),
                "payment_mode": data.payment_mode,
                "amount": CAMP_PRICE,
                "center_label": existing.get("center_label"),
                "batch_dates": existing.get("batch_dates"),
                "message": "Already registered",
                "is_duplicate": True,
            }
    # ─────────────────────────────────────────────────────────────────────────

    new_status = "payment_offline" if data.payment_mode == "cash" else "lead"
    await db.summer_camp_bookings.update_one(
        {"id": booking_id},
        {"$set": {
            "child_name": data.child_name,
            "parent_name": data.parent_name or "",
            "parent_email": data.parent_email,
            "payment_mode": data.payment_mode,
            "crm_status": new_status,
            "return_count": 1,
            "updated_at": datetime.now(timezone.utc).isoformat(),
        }}
    )

    # For cash bookings: increment tracking conversion + send WhatsApp
    if data.payment_mode == "cash":
        # Increment tracking link conversion (cash bypasses Cashfree webhook)
        source_ref = booking.get("source_ref") or booking.get("ref")
        if source_ref:
            await _increment_tracking(source_ref, "conversions")

        try:
            from .notifications import send_whatsapp_notification
            phone = booking.get("parent_phone", "")
            first_name = (data.child_name or "").split()[0] if data.child_name else "there"
            params = await _build_enrolled_wa_params(booking)
            await send_whatsapp_notification(
                phone=phone,
                template_key="summercamp_enrolled",
                params=params,
                user_name=first_name,
            )
            logging.info(f"[WA] Enrollment message sent for cash booking {booking_id}")
        except Exception as e:
            logging.warning(f"[WA] Enrollment message failed for {booking_id}: {e}")

    return {
        "booking_id": booking_id,
        "booking_ref": booking.get("booking_ref"),
        "payment_mode": data.payment_mode,
        "amount": CAMP_PRICE,
        "center_label": booking.get("center_label"),
        "batch_dates": booking.get("batch_dates"),
        "message": "Lead updated successfully",
        "is_duplicate": False,
    }


@router.post("/summer-camp/register")
async def register_summer_camp(data: SummerCampRegistration):
    """Register a summer camp booking — creates a lead record."""
    if data.age_group not in AGE_GROUPS:
        raise HTTPException(status_code=400, detail="Invalid age group")
    if data.batch_type not in ("weekday",):
        raise HTTPException(status_code=400, detail="Invalid batch type")
    if data.batch_week not in BATCH_DATES:
        raise HTTPException(status_code=400, detail="Invalid batch week")
    if data.mode not in ("offline", "online"):
        raise HTTPException(status_code=400, detail="Invalid mode")

    # Generate short booking ref
    ref_num = str(int(time.time()))[-6:]
    booking_ref = f"SC2026-{ref_num}"
    booking_id = str(uuid.uuid4())

    batch_dates = BATCH_DATES[data.batch_week][data.batch_type]
    center_label = await _resolve_center_label(data.center)

    doc = {
        "id": booking_id,
        "booking_ref": booking_ref,
        "child_name": data.child_name,
        "parent_name": data.parent_name,
        "parent_phone": data.parent_phone,
        "parent_email": data.parent_email,
        "age_group": data.age_group,
        "age_group_label": AGE_GROUPS[data.age_group]["label"],
        "age_group_ages": AGE_GROUPS[data.age_group]["ages"],
        "batch_type": data.batch_type,
        "batch_week": data.batch_week,
        "batch_dates": batch_dates,
        "mode": data.mode,
        "center": data.center,
        "center_label": center_label,
        "payment_mode": data.payment_mode,
        "amount": CAMP_PRICE,
        "payment_status": "pending",
        "crm_status": "payment_offline" if data.payment_mode == "cash" else "lead",
        "created_at": datetime.now(timezone.utc).isoformat(),
        "updated_at": datetime.now(timezone.utc).isoformat(),
    }

    await db.summer_camp_bookings.insert_one(doc)
    logging.info(f"Summer camp lead created: {booking_ref} - {data.child_name}")

    # Send enrollment WhatsApp for direct cash registrations
    if data.payment_mode == "cash":
        try:
            from .notifications import send_whatsapp_notification
            first_name = (data.child_name or "").split()[0] if data.child_name else "there"
            # Build a minimal booking dict from the submitted data for param lookup
            booking_dict = {
                "child_name": data.child_name,
                "batch_week": data.batch_week,
                "batch_dates": getattr(data, "batch_dates", None),
                "center": data.center,
                "center_label": getattr(data, "center_label", None),
            }
            params = await _build_enrolled_wa_params(booking_dict)
            await send_whatsapp_notification(
                phone=data.parent_phone,
                template_key="summercamp_enrolled",
                params=params,
                user_name=first_name,
            )
            logging.info(f"[WA] Enrollment message sent for direct cash registration {booking_ref}")
        except Exception as e:
            logging.warning(f"[WA] Enrollment message failed for {booking_ref}: {e}")

    return {
        "booking_id": booking_id,
        "booking_ref": booking_ref,
        "payment_mode": data.payment_mode,
        "amount": CAMP_PRICE,
        "center_label": center_label,
        "batch_dates": batch_dates,
        "message": "Registration successful",
    }


@router.post("/summer-camp/initiate-payment")
async def initiate_payment(data: PaymentInitRequest):
    """Create a Cashfree payment order for a summer camp booking."""
    booking = await db.summer_camp_bookings.find_one({"id": data.booking_id}, {"_id": 0})
    if not booking:
        raise HTTPException(status_code=404, detail="Booking not found")

    if booking.get("payment_status") == "paid":
        raise HTTPException(status_code=400, detail="Already paid")

    order_id = f"SC2026-{booking['id'][:8]}-{int(time.time())}"
    frontend_url = data.frontend_url or os.getenv("FRONTEND_URL", "https://oll.co")

    parent_name = (booking.get("parent_name") or "").strip()
    child_name = (booking.get("child_name") or "").strip()
    # Use child_name as fallback; ensure minimum 3 non-space chars for Cashfree
    cf_name = parent_name or f"Parent {child_name}".strip() or "OLL Parent"
    if len(cf_name.replace(" ", "")) < 3:
        cf_name = "OLL Parent"

    # Sanitize phone: Cashfree requires exactly 10 digits (no country code, no spaces/dashes)
    raw_phone = booking.get("parent_phone") or "9999999999"
    cf_phone = re.sub(r'\D', '', raw_phone)  # strip all non-digit chars
    if len(cf_phone) > 10:
        cf_phone = cf_phone[-10:]            # strip leading country code (91/+91)
    if len(cf_phone) < 10:
        cf_phone = "9999999999"              # fallback for malformed numbers

    try:
        customer = CashfreeCustomerDetails(
            customer_id=booking["id"][:50],
            customer_name=cf_name,
            customer_email=booking.get("parent_email") or f"{booking['id'][:8]}@oll.co",
            customer_phone=cf_phone,
        )
        order_meta = OrderMeta(
            return_url=f"{frontend_url}/summer-camp/success?order_id={order_id}&booking_id={booking['id']}",
            notify_url=f"{frontend_url}/api/summer-camp/webhook",
        )
        create_order_request = CreateOrderRequest(
            order_id=order_id,  # Pass our order_id so Cashfree can be queried by it later
            order_amount=CAMP_PRICE,
            order_currency="INR",
            customer_details=customer,
            order_meta=order_meta,
            order_note=f"Future Skills Summer Camp 2026 - {booking.get('age_group_label', '')} - {booking.get('child_name', '')}",
        )
        cf = get_cf_client()
        api_response = await asyncio.to_thread(cf.PGCreateOrder, CASHFREE_API_VERSION, create_order_request, None, None)

        if not api_response.data:
            raise HTTPException(status_code=500, detail="Payment gateway error")

        payment_session_id = api_response.data.payment_session_id
        payment_link = f"https://payments.cashfree.com/forms/{payment_session_id}"

        await db.summer_camp_bookings.update_one(
            {"id": booking["id"]},
            {"$set": {
                "order_id": order_id,
                "cf_order_id": str(api_response.data.cf_order_id),
                "payment_session_id": payment_session_id,
                "payment_link": payment_link,
                "updated_at": datetime.now(timezone.utc).isoformat(),
            }}
        )

        return {
            "order_id": order_id,
            "payment_session_id": payment_session_id,
            "payment_link": payment_link,
        }

    except HTTPException:
        raise
    except Exception as e:
        err_msg = str(e)
        logging.error(f"Summer camp payment initiation error for booking {data.booking_id}: {err_msg}", exc_info=True)
        # Surface the actual Cashfree error details instead of generic message
        raise HTTPException(status_code=500, detail=f"Payment initiation failed: {err_msg}")


@router.get("/summer-camp/verify/{booking_id}")
async def verify_payment(booking_id: str):
    """Check payment status for a booking. Called by success page after Cashfree redirect."""
    booking = await db.summer_camp_bookings.find_one({"id": booking_id}, {"_id": 0})
    if not booking:
        raise HTTPException(status_code=404, detail="Booking not found")

    # Already converted — return immediately without re-querying Cashfree
    if booking.get("crm_status") == "converted":
        return {"status": "PAID", "booking": booking}

    order_id = booking.get("order_id")
    if not order_id:
        return {"status": "pending", "booking": booking}

    try:
        cf = get_cf_client()
        resp = await asyncio.to_thread(cf.PGFetchOrder, CASHFREE_API_VERSION, order_id, None)
        if resp.data:
            new_status = resp.data.order_status
            if new_status == "PAID" and booking.get("crm_status") != "converted":
                await db.summer_camp_bookings.update_one(
                    {"id": booking_id},
                    {"$set": {
                        "payment_status": "paid",
                        "crm_status": "converted",
                        "updated_at": datetime.now(timezone.utc).isoformat(),
                    }}
                )
                if booking.get("source_ref"):
                    await _increment_tracking(booking["source_ref"], "conversions")
                booking["payment_status"] = "paid"
                booking["crm_status"] = "converted"

                # Fire enrollment WhatsApp message
                try:
                    from .notifications import send_whatsapp_notification
                    phone = booking.get("parent_phone", "")
                    first_name = (booking.get("child_name") or "").split()[0] or "there"
                    params = await _build_enrolled_wa_params(booking)
                    await send_whatsapp_notification(
                        phone=phone,
                        template_key="summercamp_enrolled",
                        params=params,
                        user_name=first_name,
                    )
                except Exception as wa_err:
                    logging.warning(f"[WA] Enrollment WA failed in verify: {wa_err}")

            return {"status": new_status, "booking": booking}
    except Exception as e:
        logging.warning(f"Summer camp verify error: {e}")

    return {"status": booking.get("payment_status", "pending"), "booking": booking}


@router.post("/summer-camp/webhook")
async def summer_camp_webhook(request: Request):
    """Cashfree webhook for summer camp payments."""
    try:
        payload = await request.json()
        order_data = payload.get("data", {}).get("order", {})
        order_id = order_data.get("order_id")
        order_status = order_data.get("order_status")

        if not order_id:
            return {"status": "ignored"}

        if order_status == "PAID":
            booking = await db.summer_camp_bookings.find_one({"order_id": order_id}, {"_id": 0})
            if booking:
                await db.summer_camp_bookings.update_one(
                    {"order_id": order_id},
                    {"$set": {
                        "payment_status": "paid",
                        "crm_status": "converted",
                        "updated_at": datetime.now(timezone.utc).isoformat(),
                    }}
                )
                # Increment conversion on tracking link
                if booking.get("source_ref"):
                    await _increment_tracking(booking["source_ref"], "conversions")

                # Send enrollment WhatsApp confirmation
                try:
                    from .notifications import send_whatsapp_notification
                    phone = booking.get("parent_phone", "")
                    first_name = (booking.get("child_name") or "").split()[0] or "there"
                    params = await _build_enrolled_wa_params(booking)
                    await send_whatsapp_notification(
                        phone=phone,
                        template_key="summercamp_enrolled",
                        params=params,
                        user_name=first_name,
                    )
                    logging.info(f"[WA] Enrollment message sent for online booking {order_id}")
                except Exception as e:
                    logging.warning(f"[WA] Enrollment message failed for {order_id}: {e}")

                logging.info(f"Summer camp payment confirmed via webhook: {order_id}")

        return {"status": "ok"}
    except Exception as e:
        logging.error(f"Summer camp webhook error: {e}")
        return {"status": "error"}


@router.get("/summer-camp/availability")
async def get_batch_availability(age_group: str, center: str):
    """Public: return spots_left per batch week for a given age_group + center combo."""
    SPOTS_PER_BATCH = 10
    result = {}
    for wk in ["week1", "week2", "week3", "week4"]:
        # Only count CONFIRMED students (online paid + cash at center)
        # Leads and phone-captures should NOT block spots
        count = await db.summer_camp_bookings.count_documents({
            "batch_week": wk,
            "age_group": age_group,
            "center": center,
            "crm_status": {"$in": ["converted", "payment_offline"]},
        })
        spots_left = max(0, SPOTS_PER_BATCH - int(count))
        result[wk] = {
            "booked": int(count),
            "spots_left": spots_left,
            "spots_total": SPOTS_PER_BATCH,
            "full": spots_left == 0,
        }
    return result


@router.get("/summer-camp/bookings")
async def get_summer_camp_bookings(
    crm_status: Optional[str] = None,
    age_group: Optional[str] = None,
    user: dict = Depends(get_current_user)
):
    """Admin: list all summer camp bookings."""
    query = {}
    if crm_status:
        query["crm_status"] = crm_status
    if age_group:
        query["age_group"] = age_group

    bookings = await db.summer_camp_bookings.find(query, {"_id": 0}).sort("created_at", -1).to_list(500)
    return bookings


@router.get("/summer-camp/stats")
async def get_summer_camp_stats(user: dict = Depends(get_current_user)):
    """Admin stats for summer camp."""
    total = await db.summer_camp_bookings.count_documents({})
    converted = await db.summer_camp_bookings.count_documents({"crm_status": "converted"})
    leads = await db.summer_camp_bookings.count_documents({"crm_status": "lead"})
    return {"total": total, "converted": converted, "leads": leads}


# ── CRM Management Endpoints ───────────────────────────────────────────────────

@router.patch("/summer-camp/bookings/{booking_id}")
async def update_booking(
    booking_id: str,
    data: BookingUpdate,
    user: dict = Depends(get_current_user)
):
    """Admin: edit booking details."""
    booking = await db.summer_camp_bookings.find_one({"id": booking_id})
    if not booking:
        raise HTTPException(status_code=404, detail="Booking not found")

    update_fields = {"updated_at": datetime.now(timezone.utc).isoformat()}
    if data.child_name is not None:
        update_fields["child_name"] = data.child_name
    if data.parent_name is not None:
        update_fields["parent_name"] = data.parent_name
    if data.parent_phone is not None:
        update_fields["parent_phone"] = data.parent_phone
    if data.parent_email is not None:
        update_fields["parent_email"] = data.parent_email

    await db.summer_camp_bookings.update_one({"id": booking_id}, {"$set": update_fields})
    updated = await db.summer_camp_bookings.find_one({"id": booking_id}, {"_id": 0})
    return updated


@router.delete("/summer-camp/bookings/{booking_id}")
async def delete_booking(booking_id: str, user: dict = Depends(get_current_user)):
    """Admin: delete a booking."""
    result = await db.summer_camp_bookings.delete_one({"id": booking_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Booking not found")
    return {"message": "Booking deleted"}


@router.patch("/summer-camp/bookings/{booking_id}/status")
async def update_booking_status(
    booking_id: str,
    data: StatusUpdate,
    user: dict = Depends(get_current_user)
):
    """Admin: update CRM status (phone_captured | lead | hot_lead | converted | payment_offline | lost_lead)."""
    valid = {"phone_captured", "lead", "hot_lead", "converted", "payment_offline", "lost_lead"}
    if data.crm_status not in valid:
        raise HTTPException(status_code=400, detail=f"Invalid status. Must be one of: {valid}")

    update_fields: dict = {
        "crm_status": data.crm_status,
        "updated_at": datetime.now(timezone.utc).isoformat(),
    }
    if data.crm_status == "lost_lead" and data.lost_reason:
        update_fields["lost_reason"] = data.lost_reason

    result = await db.summer_camp_bookings.update_one(
        {"id": booking_id},
        {"$set": update_fields}
    )
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Booking not found")
    return {"success": True, "crm_status": data.crm_status}


@router.post("/summer-camp/bookings/{booking_id}/comment")
async def add_booking_comment(
    booking_id: str,
    data: CommentAdd,
    user: dict = Depends(get_current_user)
):
    """Admin: add a comment or call log entry to a booking."""
    comment = {
        "id": str(uuid.uuid4()),
        "text": data.text,
        "author": data.author or user.get("name", user.get("email", "Admin")),
        "comment_type": data.comment_type or "comment",
        "created_at": datetime.now(timezone.utc).isoformat(),
    }
    if data.comment_type == "call_done":
        comment["call_date"] = data.call_date or datetime.now(timezone.utc).strftime("%Y-%m-%d")
        comment["call_time"] = data.call_time or datetime.now(timezone.utc).strftime("%H:%M")
    result = await db.summer_camp_bookings.update_one(
        {"id": booking_id},
        {"$push": {"comments": comment}}
    )
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Booking not found")
    return comment


@router.patch("/summer-camp/bookings/{booking_id}/followup-status")
async def update_followup_status(
    booking_id: str,
    data: FollowupStatusUpdate,
    user: dict = Depends(get_current_user)
):
    """Admin: update the CRM follow-up status for a booking."""
    update_fields = {
        "followup_status": data.followup_status,
        "followup_status_updated_at": datetime.now(timezone.utc).isoformat(),
        "followup_status_updated_by": user.get("name", user.get("email", "Admin")),
    }
    if data.followup_status == "callback_requested" and data.callback_date:
        update_fields["callback_date"] = data.callback_date
        update_fields["callback_time"] = data.callback_time or ""
    result = await db.summer_camp_bookings.update_one(
        {"id": booking_id},
        {"$set": update_fields}
    )
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Booking not found")
    return {"success": True, "followup_status": data.followup_status}


@router.api_route("/summer-camp/brochure", methods=["GET", "HEAD"])
async def serve_summer_camp_brochure():
    """Publicly accessible Summer Camp PDF brochure — used as WhatsApp attachment."""
    pdf_path = os.path.join(os.path.dirname(__file__), "..", "static", "summer-camp-brochure.pdf")
    pdf_path = os.path.abspath(pdf_path)
    if not os.path.exists(pdf_path):
        raise HTTPException(status_code=404, detail="Brochure not found")
    return FileResponse(pdf_path, media_type="application/pdf", filename="Summer-Camp-Brochure-2026.pdf")



@router.get("/summer-camp/dashboard")
async def get_summer_camp_dashboard(user: dict = Depends(get_current_user)):
    """Admin dashboard: week × age group × center breakdown, 10 spots per group."""
    SPOTS_PER_BATCH = 10

    AGE_GROUPS = [
        {"key": "explorers",  "label": "Little Explorers",  "ages": "4–8"},
        {"key": "creators",   "label": "Tech Creators",     "ages": "9–12"},
        {"key": "innovators", "label": "Future Innovators", "ages": "13–16"},
    ]

    WEEK_DATES = {
        "week1": "May 4–8, 2026",
        "week2": "May 11–15, 2026",
        "week3": "May 18–22, 2026",
        "week4": "May 25–29, 2026",
    }

    bookings = await db.summer_camp_bookings.find({}, {"_id": 0}).to_list(2000)

    # Build nested structure: week → age_group → {stats, by_center}
    data = {}
    for wk, wk_dates in WEEK_DATES.items():
        data[wk] = {
            "week": wk,
            "dates": wk_dates,
            "week_label": f"Week {wk[-1]}",
            "age_groups": {
                ag["key"]: {
                    "label": ag["label"],
                    "ages": ag["ages"],
                    "total": 0,
                    "converted": 0,
                    "leads": 0,
                    "phone_captured": 0,
                    "lost": 0,
                    "spots_total": SPOTS_PER_BATCH,
                    "by_center": {},
                }
                for ag in AGE_GROUPS
            },
        }

    for b in bookings:
        wk = b.get("batch_week")
        ag = b.get("age_group")
        if wk not in data or ag not in data[wk]["age_groups"]:
            continue

        cell = data[wk]["age_groups"][ag]
        cell["total"] += 1
        status = b.get("crm_status", "")
        if status in ("converted", "payment_offline"):
            cell["converted"] += 1
        elif status == "hot_lead":
            cell["leads"] += 1
        elif status == "lead":
            cell["leads"] += 1
        elif status == "phone_captured":
            cell["phone_captured"] += 1
        elif status == "lost_lead":
            cell["lost"] += 1

        # Center tracking (track full stats per center, not just count) — normalize label
        center = _normalize_center_display(b.get("center_label") or b.get("center") or "Unknown")
        if center not in cell["by_center"]:
            cell["by_center"][center] = {
                "total": 0, "converted": 0, "leads": 0,
                "phone_captured": 0, "lost": 0,
            }
        by_c = cell["by_center"][center]
        by_c["total"] += 1
        if status in ("converted", "payment_offline"):
            by_c["converted"] += 1
        elif status in ("lead", "hot_lead"):
            by_c["leads"] += 1
        elif status == "phone_captured":
            by_c["phone_captured"] += 1
        elif status == "lost_lead":
            by_c["lost"] += 1

    # Compute spots_left for each cell (aggregate) + per-center
    for wk_data in data.values():
        for ag_data in wk_data["age_groups"].values():
            ag_data["spots_left"] = max(0, SPOTS_PER_BATCH - ag_data["converted"])
            for by_c in ag_data["by_center"].values():
                by_c["spots_left"] = max(0, SPOTS_PER_BATCH - by_c["converted"])

    # Summary totals
    total_bookings = len(bookings)
    total_registrations = sum(1 for b in bookings)  # all bookings = registrations
    total_hot_leads = sum(1 for b in bookings if b.get("crm_status") == "hot_lead")
    total_leads = sum(1 for b in bookings if b.get("crm_status") == "lead")
    total_converted_online = sum(1 for b in bookings if b.get("crm_status") == "converted")
    total_converted_offline = sum(1 for b in bookings if b.get("crm_status") == "payment_offline")
    total_converted = total_converted_online + total_converted_offline
    total_revenue = total_converted_online * CAMP_PRICE  # Only online payments have actual revenue

    # Conversion funnel ratios
    reg_to_hot = round((total_hot_leads / total_registrations * 100), 1) if total_registrations else 0
    hot_to_conv = round((total_converted / max(total_hot_leads, 1) * 100), 1) if total_hot_leads else 0
    reg_to_conv = round((total_converted / total_registrations * 100), 1) if total_registrations else 0

    # Collect unique centers (normalized to canonical names)
    all_centers = sorted(set(
        _normalize_center_display(b.get("center_label") or b.get("center") or "")
        for b in bookings
        if b.get("center_label") or b.get("center")
    ))

    # Followup status breakdown
    followup_counts = {
        "not_contacted": 0,
        "call_not_picked": 0,
        "call_cut": 0,
        "callback_requested": 0,
        "interested": 0,
        "non_serviceable": 0,
    }
    for b in bookings:
        fs = b.get("followup_status") or "not_contacted"
        if fs in followup_counts:
            followup_counts[fs] += 1
        else:
            followup_counts["not_contacted"] += 1

    return {
        "total_bookings": total_bookings,
        "total_revenue": total_revenue,
        "converted": total_converted,
        "converted_online": total_converted_online,
        "converted_offline": total_converted_offline,
        "hot_leads": total_hot_leads,
        "leads": total_leads,
        "followup_status": followup_counts,
        "funnel": {
            "registrations": total_registrations,
            "hot_leads": total_hot_leads,
            "converted": total_converted,
            "reg_to_hot_pct": reg_to_hot,
            "hot_to_conv_pct": hot_to_conv,
            "reg_to_conv_pct": reg_to_conv,
        },
        "centers": all_centers,
        "weeks": sorted(data.values(), key=lambda x: x["week"]),
        # Keep old age_summary for the revenue bar chart
        "age_summary": [
            {
                "age_group": ag["key"],
                "label": ag["label"],
                "total": sum(1 for b in bookings if b.get("age_group") == ag["key"]),
                "converted": sum(1 for b in bookings if b.get("age_group") == ag["key"] and b.get("crm_status") in ("converted", "payment_offline")),
                "revenue": sum(CAMP_PRICE for b in bookings if b.get("age_group") == ag["key"] and b.get("crm_status") == "converted"),
            }
            for ag in AGE_GROUPS
        ],
    }


# ─── Bulk Import ───────────────────────────────────────────────────────────────

VALID_AGE_GROUPS = {"explorers", "creators", "innovators"}
VALID_STATUSES = {"phone_captured", "lead", "hot_lead", "converted", "payment_offline", "lost_lead"}
VALID_BATCH_WEEKS = {"week1", "week2", "week3", "week4"}

def _normalize_center_display(label: str) -> str:
    """Normalize messy historical center_label variants to canonical display names."""
    if not label:
        return "Unknown"
    low = label.lower().strip()
    # Strip parenthesized address suffixes, e.g. "Andheri West (Azad Nagar, Mumbai)"
    # Keep only the first part before "("
    core = low.split("(")[0].strip()
    if "andheri" in core or "azad nagar" in core:
        return "OLL Andheri Center"
    if "mira road" in core or "mira-road" in core:
        return "OLL Mira Road Center"
    if "borivali" in core or "kandivali" in core:
        return "OLL Borivali Center"
    if "malad" in core:
        return "OLL Malad Center"
    if "thane" in core:
        return "OLL Thane Center"
    if "online" in core:
        return "Online"
    # Fall back to original with parenthesis stripped
    return label.split("(")[0].strip() or label
    val = (val or "").strip().lower()
    if val in VALID_AGE_GROUPS:
        return val
    if val in ("4-8", "4–8", "littleexplorers", "little explorers", "explorers"):
        return "explorers"
    if val in ("9-12", "9–12", "techcreators", "tech creators", "creators"):
        return "creators"
    if val in ("13-16", "13–16", "futureinnovators", "future innovators", "innovators"):
        return "innovators"
    return ""

def _normalize_batch_week(val: str) -> str:
    val = (val or "").strip().lower().replace(" ", "")
    if val in VALID_BATCH_WEEKS:
        return val
    mapping = {
        "batch1": "week1", "batch2": "week2", "batch3": "week3", "batch4": "week4",
        "1": "week1", "2": "week2", "3": "week3", "4": "week4",
        "may4-8": "week1", "may11-15": "week2", "may18-22": "week3", "may25-29": "week4",
    }
    return mapping.get(val, "week1")

def _normalize_status(val: str) -> str:
    val = (val or "").strip().lower().replace(" ", "_")
    if val in VALID_STATUSES:
        return val
    return "lead"


async def _send_import_wa(booking: dict) -> None:
    """Fire the appropriate WhatsApp notification for a freshly-imported booking."""
    from .notifications import send_whatsapp_notification
    phone = booking.get("parent_phone", "")
    if not phone:
        return
    status = booking.get("crm_status", "lead")
    child_name = booking.get("child_name") or booking.get("parent_name") or "Student"
    first_name = child_name.strip().split()[0] if child_name.strip() else "Student"

    try:
        if status == "phone_captured":
            await send_whatsapp_notification(
                phone=phone,
                template_key="summercamp_followup",
                params=[],
                user_name=first_name,
                media={"url": SUMMER_CAMP_BROCHURE_URL, "filename": SUMMER_CAMP_BROCHURE_FILENAME},
            )
        elif status in ("lead", "hot_lead"):
            await send_whatsapp_notification(
                phone=phone,
                template_key="summercamp_payment_pending",
                params=["$FirstName"],
                user_name=first_name,
                media={"url": SUMMER_CAMP_BROCHURE_URL, "filename": SUMMER_CAMP_BROCHURE_FILENAME},
            )
        elif status in ("converted", "payment_offline"):
            params = await _build_enrolled_wa_params(booking)
            await send_whatsapp_notification(
                phone=phone,
                template_key="summercamp_enrolled",
                params=params,
                user_name=first_name,
            )
        # lost_lead → no notification
    except Exception as e:
        logging.warning(f"[BulkImport] WA notification failed for {phone}: {e}")


@router.get("/summer-camp/bulk-import-sample")
async def get_bulk_import_sample(user: dict = Depends(get_current_user)):
    """Return a sample XLSX file that shows the expected import format."""
    if not OPENPYXL_AVAILABLE:
        raise HTTPException(status_code=500, detail="openpyxl not installed")

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Summer Camp Leads"

    headers = [
        "parent_phone", "child_name", "parent_name", "parent_email",
        "age_group", "batch_week", "center_name", "crm_status",
    ]

    # Style header row (no notes/description row — it caused import errors)
    header_fill = PatternFill("solid", fgColor="1E3A5F")
    header_font = Font(color="FFFFFF", bold=True)
    col_widths  = [18, 20, 20, 28, 14, 12, 28, 20]
    for col_idx, (h, w) in enumerate(zip(headers, col_widths), start=1):
        cell = ws.cell(row=1, column=col_idx, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")
        ws.column_dimensions[cell.column_letter].width = w

    # Sample rows — age_group uses human-readable age ranges
    samples = [
        ["+919876543210", "Aryan Sharma", "Raj Sharma", "raj@gmail.com", "9-12", "week1", "OLL Andheri Center", "lead"],
        ["9123456789",    "Priya Patel",  "Sunita Patel", "sunita@gmail.com", "4-8",  "week2", "",                  "phone_captured"],
        ["8800001234",    "",             "Neha Singh",  "",               "13-16", "",      "",                  "lead"],
    ]
    for row_idx, row in enumerate(samples, start=2):
        for col_idx, val in enumerate(row, start=1):
            ws.cell(row=row_idx, column=col_idx, value=val)

    # Legend row with light background
    legend = [
        "With or without +91",
        "Optional",
        "Optional",
        "Optional",
        "4-8 / 9-12 / 13-16",
        "week1/week2/week3/week4 (optional)",
        "Optional",
        "lead/hot_lead/phone_captured/converted/payment_offline/lost_lead",
    ]
    legend_fill = PatternFill("solid", fgColor="EFF6FF")
    legend_font = Font(italic=True, color="6B7280", size=9)
    for col_idx, note in enumerate(legend, start=1):
        cell = ws.cell(row=len(samples) + 2, column=col_idx, value=note)
        cell.fill = legend_fill
        cell.font = legend_font

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=SummerCamp_Import_Sample.xlsx"},
    )




class AddIndividualLeadRequest(BaseModel):
    child_name: str
    parent_name: Optional[str] = ""
    parent_phone: str
    parent_email: Optional[str] = ""
    age_group: Optional[str] = ""    # explorers | creators | innovators
    batch_week: Optional[str] = ""   # week1-week4
    mode: Optional[str] = "offline"  # offline | online
    center: Optional[str] = ""
    payment_mode: str = "cash"       # cash | cashfree
    source: Optional[str] = "admin"  # tracking source
    send_notifications: bool = True  # whether to fire WA + email


@router.post("/summer-camp/add-lead")
async def add_individual_lead(
    background_tasks: BackgroundTasks,
    data: AddIndividualLeadRequest,
    user: dict = Depends(get_current_user),
):
    """Admin endpoint to add a single Summer Camp lead with optional WA + email notifications."""
    # Sanitize phone
    phone = re.sub(r"[^0-9+]", "", data.parent_phone or "")
    if phone.startswith("91") and len(phone) == 12:
        phone = "+" + phone
    elif not phone.startswith("+"):
        phone = "+91" + phone.lstrip("0")

    ref_num = str(int(time.time()))[-6:]
    booking_ref = f"SC2026-{ref_num}"
    booking_id = str(uuid.uuid4())

    batch_dates = ""
    if data.batch_week and data.batch_week in BATCH_DATES:
        batch_dates = BATCH_DATES[data.batch_week].get("weekday", "")

    center_label = await _resolve_center_label(data.center) if data.center else ""

    doc = {
        "id": booking_id,
        "booking_ref": booking_ref,
        "child_name": data.child_name,
        "parent_name": data.parent_name or "",
        "parent_phone": phone,
        "parent_email": data.parent_email or "",
        "age_group": data.age_group or "",
        "age_group_label": AGE_GROUPS.get(data.age_group or "", {}).get("label", ""),
        "age_group_ages": AGE_GROUPS.get(data.age_group or "", {}).get("ages", ""),
        "batch_type": "weekday",
        "batch_week": data.batch_week or "",
        "batch_dates": batch_dates,
        "mode": data.mode or "offline",
        "center": data.center or "",
        "center_label": center_label,
        "payment_mode": data.payment_mode,
        "amount": CAMP_PRICE,
        "payment_status": "pending",
        "crm_status": "payment_offline" if data.payment_mode == "cash" else "lead",
        "source": data.source or "admin",
        "added_by": user.get("name") or user.get("email", "Admin"),
        "created_at": datetime.now(timezone.utc).isoformat(),
        "updated_at": datetime.now(timezone.utc).isoformat(),
    }

    await db.summer_camp_bookings.insert_one(doc)
    logging.info(f"[Admin] Individual lead added: {booking_ref} — {data.child_name} by {doc['added_by']}")

    if data.send_notifications:
        async def _fire_notifications(d: dict):
            first_name = (d.get("child_name") or "").split()[0] or "there"
            # ── WhatsApp ─────────────────────────────────────────────────────
            try:
                from .notifications import send_whatsapp_notification
                if d.get("batch_week"):
                    # Enrolled — send enrolled template
                    params = await _build_enrolled_wa_params(d)
                    tpl = "summercamp_enrolled"
                else:
                    # No batch yet — send followup/interest template
                    params = [first_name]
                    tpl = "summercamp_followup"
                await send_whatsapp_notification(
                    phone=d["parent_phone"],
                    template_key=tpl,
                    params=params,
                    user_name=first_name,
                )
                logging.info(f"[Admin Lead WA] Sent {tpl} to {d['parent_phone']}")
            except Exception as e:
                logging.warning(f"[Admin Lead WA] Failed: {e}")

            # ── Email ─────────────────────────────────────────────────────────
            try:
                if d.get("parent_email"):
                    from .shared import get_resend_api_key, SENDER_EMAIL
                    import resend as _resend
                    key = await get_resend_api_key()
                    if key:
                        _resend.api_key = key
                        child = d.get("child_name", "your child")
                        batch_info = d.get("batch_dates") or d.get("batch_week") or "TBD"
                        center_info = d.get("center_label") or d.get("center") or "TBD"
                        body = f"""<div style="font-family:Arial,sans-serif;max-width:600px;margin:0 auto">
<h2 style="color:#f97316">OLL Summer Camp 2026 — Registration Confirmed</h2>
<p>Hi {d.get('parent_name') or 'there'},</p>
<p>We've registered <strong>{child}</strong> for the OLL AI Summer Camp 2026!</p>
<table style="width:100%;border-collapse:collapse;margin:16px 0">
  <tr><td style="padding:8px;border:1px solid #e5e7eb;font-weight:bold;background:#fef9f0">Booking Ref</td><td style="padding:8px;border:1px solid #e5e7eb">{d.get('booking_ref','')}</td></tr>
  <tr><td style="padding:8px;border:1px solid #e5e7eb;font-weight:bold;background:#fef9f0">Batch</td><td style="padding:8px;border:1px solid #e5e7eb">{batch_info}</td></tr>
  <tr><td style="padding:8px;border:1px solid #e5e7eb;font-weight:bold;background:#fef9f0">Center</td><td style="padding:8px;border:1px solid #e5e7eb">{center_info}</td></tr>
  <tr><td style="padding:8px;border:1px solid #e5e7eb;font-weight:bold;background:#fef9f0">Age Group</td><td style="padding:8px;border:1px solid #e5e7eb">{d.get('age_group_label') or d.get('age_group','')}</td></tr>
  <tr><td style="padding:8px;border:1px solid #e5e7eb;font-weight:bold;background:#fef9f0">Amount</td><td style="padding:8px;border:1px solid #e5e7eb">₹{int(d.get('amount',1999))}</td></tr>
</table>
<p>If you have any questions, reply to this email or WhatsApp us.</p>
<p style="color:#6b7280;font-size:12px">— OLL Team</p>
</div>"""
                        import asyncio as _asyncio
                        await _asyncio.to_thread(_resend.Emails.send, {
                            "from": SENDER_EMAIL,
                            "to": [d["parent_email"]],
                            "subject": f"OLL Summer Camp 2026 — {child} is Registered! ({d.get('booking_ref','')})",
                            "html": body,
                        })
                        logging.info(f"[Admin Lead Email] Sent to {d['parent_email']}")
            except Exception as e:
                logging.warning(f"[Admin Lead Email] Failed: {e}")

        background_tasks.add_task(_fire_notifications, doc)

    return {
        "success": True,
        "booking_id": booking_id,
        "booking_ref": booking_ref,
        "message": f"Lead added successfully. {'Notifications queued.' if data.send_notifications else 'No notifications sent.'}",
    }


@router.post("/summer-camp/bulk-import")
async def bulk_import_leads(
    background_tasks: BackgroundTasks,
    file: UploadFile = File(...),
    user: dict = Depends(get_current_user),
):
    """
    Bulk import Summer Camp leads from XLSX.
    - Skips rows where parent_phone already exists in the DB.
    - Fires WhatsApp notification (based on crm_status) immediately for each imported lead.
    """
    if not OPENPYXL_AVAILABLE:
        raise HTTPException(status_code=500, detail="openpyxl not installed on server")
    if not file.filename.endswith((".xlsx", ".xls")):
        raise HTTPException(status_code=400, detail="Only .xlsx files are supported")

    content = await file.read()
    try:
        wb = openpyxl.load_workbook(io.BytesIO(content), data_only=True)
    except Exception:
        raise HTTPException(status_code=400, detail="Could not parse file. Please use the sample template.")

    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        raise HTTPException(status_code=400, detail="File is empty")

    # First non-empty row = headers
    header_row = rows[0]
    headers_raw = [str(h).strip().lower().replace(" ", "_") if h else "" for h in header_row]

    def col(name):
        try:
            return headers_raw.index(name)
        except ValueError:
            return -1

    ph_col = col("parent_phone")
    if ph_col == -1:
        raise HTTPException(status_code=400, detail="Column 'parent_phone' is required")

    imported, skipped, errors = [], [], []

    for row_num, row in enumerate(rows[1:], start=2):
        if not any(row):
            continue  # skip fully empty rows

        def get(c):
            if c == -1 or c >= len(row):
                return ""
            val = row[c]
            return str(val).strip() if val is not None else ""

        raw_phone_val = get(ph_col)

        # Skip rows that look like header/notes rows (first char is a letter, not a digit/+)
        first_char = raw_phone_val[0] if raw_phone_val else ""
        if first_char and first_char not in "0123456789+":
            continue  # silently skip notes/description rows

        # Sanitize phone: accept +91…, 91…(12 digits), or 10 digits plain
        raw_phone = re.sub(r'\D', '', raw_phone_val)
        if len(raw_phone) == 12 and raw_phone.startswith("91"):
            raw_phone = raw_phone[2:]   # strip 91 prefix
        elif len(raw_phone) == 13 and raw_phone.startswith("091"):
            raw_phone = raw_phone[3:]
        elif len(raw_phone) > 10:
            raw_phone = raw_phone[-10:]  # fallback: take last 10 digits

        if len(raw_phone) != 10:
            if raw_phone_val:  # only error if the cell had something
                errors.append(f"Row {row_num}: invalid phone '{raw_phone_val}' (need 10 digits)")
            continue

        # Skip duplicate
        existing = await db.summer_camp_bookings.find_one({"parent_phone": raw_phone}, {"_id": 0, "id": 1})
        if existing:
            skipped.append(raw_phone)
            continue

        child_name  = get(col("child_name"))
        parent_name = get(col("parent_name"))
        parent_email = get(col("parent_email"))
        age_group   = _normalize_age_group(get(col("age_group"))) or "explorers"
        batch_week  = _normalize_batch_week(get(col("batch_week")))
        center_name = get(col("center_name"))
        crm_status  = _normalize_status(get(col("crm_status")))

        # Resolve center
        center_id = ""
        center_label = center_name
        if center_name:
            center_doc = await db.centers.find_one(
                {"name": {"$regex": center_name, "$options": "i"}},
                {"_id": 0, "id": 1, "name": 1, "area": 1, "city": 1},
            )
            if center_doc:
                center_id = center_doc.get("id", "")
                area = center_doc.get("area", "")
                city = center_doc.get("city", "")
                center_label = f"{center_doc['name']} ({area}, {city})" if area and city else center_doc["name"]

        ref_num = str(int(time.time() * 1000))[-6:]
        booking_id = str(uuid.uuid4())
        booking_ref = f"SC2026-{ref_num}"
        batch_dates = BATCH_DATES.get(batch_week, {}).get("weekday", "")

        doc = {
            "id": booking_id,
            "booking_ref": booking_ref,
            "parent_phone": raw_phone,
            "child_name": child_name,
            "parent_name": parent_name,
            "parent_email": parent_email,
            "age_group": age_group,
            "age_group_label": AGE_GROUPS.get(age_group, {}).get("label", ""),
            "age_group_ages": AGE_GROUPS.get(age_group, {}).get("ages", ""),
            "batch_type": "weekday",
            "batch_week": batch_week,
            "batch_dates": batch_dates,
            "mode": "offline",
            "center": center_id,
            "center_label": center_label,
            "payment_mode": "cashfree" if crm_status not in ("converted", "payment_offline") else "cash",
            "amount": CAMP_PRICE,
            "payment_status": "paid" if crm_status in ("converted", "payment_offline") else "pending",
            "crm_status": crm_status,
            "source_ref": "",
            "source_name": "Bulk Import",
            "imported_by": user.get("email", "admin"),
            "created_at": datetime.now(timezone.utc).isoformat(),
            "updated_at": datetime.now(timezone.utc).isoformat(),
        }
        await db.summer_camp_bookings.insert_one(doc)
        imported.append(booking_id)

        # Fire WA notification in background so import doesn't block
        background_tasks.add_task(_send_import_wa, doc)

        logging.info(f"[BulkImport] Imported {raw_phone} ({crm_status}) ref={booking_ref}")

    return {
        "imported": len(imported),
        "skipped": len(skipped),
        "errors": errors,
        "message": f"Successfully imported {len(imported)} leads. {len(skipped)} skipped (duplicate phone). {len(errors)} errors.",
    }


# ─── Summer Camp Follow-Up Constants ───────────────────────────────────────────
# PDF brochure URL — served via the backend's /api/summer-camp/brochure endpoint
# Uses BACKEND_PUBLIC_URL from env (set to the production/preview domain)
_BACKEND_PUBLIC_URL = os.environ.get("BACKEND_PUBLIC_URL", os.environ.get("FRONTEND_URL", "")).rstrip("/")
SUMMER_CAMP_BROCHURE_URL = f"{_BACKEND_PUBLIC_URL}/api/summer-camp/brochure"
SUMMER_CAMP_BROCHURE_FILENAME = "Summer Camp Brochure 2026"


def _parse_created_dt(raw):
    """Parse an ISO string or datetime to a timezone-aware datetime. Returns None on failure."""
    if isinstance(raw, str):
        return datetime.fromisoformat(raw.replace("Z", "+00:00"))
    if isinstance(raw, datetime):
        return raw if raw.tzinfo else raw.replace(tzinfo=timezone.utc)
    return None


async def check_summer_camp_followups() -> None:
    """
    Scheduled job: Send a WhatsApp follow-up (with brochure PDF) to leads who
    entered their phone number but did NOT complete registration, after 5 minutes.
    Runs every 1 minute. Marks sent leads so they never receive a duplicate.
    """
    from .notifications import send_whatsapp_notification

    five_min_ago = datetime.now(timezone.utc).timestamp() - (5 * 60)
    max_age = datetime.now(timezone.utc).timestamp() - (30 * 24 * 60 * 60)

    try:
        # Find uncompleted leads: phone captured, no follow-up sent, older than 5 min
        cursor = db.summer_camp_bookings.find(
            {
                "crm_status": "phone_captured",
                "follow_up_wa_sent": {"$ne": True},
            },
            {"_id": 0, "id": 1, "parent_phone": 1, "child_name": 1, "created_at": 1}
        )
        bookings = await cursor.to_list(length=100)

        for booking in bookings:
            created_raw = booking.get("created_at", "")
            try:
                if isinstance(created_raw, str):
                    created_dt = datetime.fromisoformat(created_raw.replace("Z", "+00:00"))
                elif isinstance(created_raw, datetime):
                    created_dt = created_raw if created_raw.tzinfo else created_raw.replace(tzinfo=timezone.utc)
                else:
                    continue

                if created_dt.timestamp() > five_min_ago:
                    continue  # Not old enough yet

                # Skip leads older than 30 days — no point messaging ancient leads
                if created_dt.timestamp() < max_age:
                    continue

            except Exception as parse_err:
                print(f"[SC Followup] Could not parse created_at for {booking.get('id')}: {parse_err}")
                continue

            phone = booking.get("parent_phone", "")
            if not phone:
                continue

            # Mark as sent FIRST to prevent duplicates if the API call hangs
            await db.summer_camp_bookings.update_one(
                {"id": booking["id"]},
                {"$set": {
                    "follow_up_wa_sent": True,
                    "follow_up_wa_sent_at": datetime.now(timezone.utc).isoformat(),
                }}
            )

            result = await send_whatsapp_notification(
                phone=phone,
                template_key="summercamp_followup",
                params=[],
                user_name=booking.get("child_name", ""),
                media={
                    "url": SUMMER_CAMP_BROCHURE_URL,
                    "filename": SUMMER_CAMP_BROCHURE_FILENAME,
                }
            )
            print(f"[SC Followup] {'Sent' if result.get('success') else 'Failed'} for {phone} (id={booking['id']})")

    except Exception as e:
        print(f"[SC Followup] Scheduler error: {e}")


async def check_summer_camp_payment_pending() -> None:
    """
    Scheduled job: Send a WhatsApp follow-up (with brochure PDF) to leads who
    filled in their details (crm_status='lead') but did NOT complete payment,
    after 5 minutes. Uses campaign 'summercamppaymentpending' with $FirstName param.
    Marks sent leads so they never receive a duplicate.
    Uses updated_at (not created_at) so the 5-min timer starts from when they
    filled their details — preventing overlap with the phone-captured follow-up.
    """
    from .notifications import send_whatsapp_notification

    five_min_ago = datetime.now(timezone.utc).timestamp() - (5 * 60)
    max_age_ts = datetime.now(timezone.utc).timestamp() - (30 * 24 * 60 * 60)

    try:
        cursor = db.summer_camp_bookings.find(
            {
                "crm_status": "lead",
                "payment_followup_wa_sent": {"$ne": True},
            },
            {"_id": 0, "id": 1, "parent_phone": 1, "child_name": 1, "parent_name": 1,
             "created_at": 1, "updated_at": 1}
        )
        bookings = await cursor.to_list(length=100)

        for booking in bookings:
            # Use updated_at (when details were filled) — prevents immediate trigger
            # for users who just moved from phone_captured → lead
            ref_raw = booking.get("updated_at") or booking.get("created_at", "")
            try:
                ref_dt = _parse_created_dt(ref_raw)
                if not ref_dt or ref_dt.timestamp() > five_min_ago:
                    continue  # Not old enough yet
                # Skip leads older than 30 days
                created_dt = _parse_created_dt(booking.get("created_at", ""))
                if created_dt and created_dt.timestamp() < max_age_ts:
                    continue
            except Exception as parse_err:
                print(f"[SC PayPending] Could not parse timestamp for {booking.get('id')}: {parse_err}")
                continue

            phone = booking.get("parent_phone", "")
            if not phone:
                continue

            # Derive first name from child_name or parent_name
            raw_name = booking.get("child_name") or booking.get("parent_name") or "Student"
            first_name = raw_name.strip().split()[0] if raw_name.strip() else "Student"

            # Mark FIRST to prevent duplicate sends
            await db.summer_camp_bookings.update_one(
                {"id": booking["id"]},
                {"$set": {
                    "payment_followup_wa_sent": True,
                    "payment_followup_wa_sent_at": datetime.now(timezone.utc).isoformat(),
                }}
            )

            result = await send_whatsapp_notification(
                phone=phone,
                template_key="summercamp_payment_pending",
                params=["$FirstName"],
                user_name=first_name,
                media={
                    "url": SUMMER_CAMP_BROCHURE_URL,
                    "filename": SUMMER_CAMP_BROCHURE_FILENAME,
                },
            )
            print(f"[SC PayPending] {'Sent' if result.get('success') else 'Failed'} for {phone} (id={booking['id']}, name={first_name})")

    except Exception as e:
        print(f"[SC PayPending] Scheduler error: {e}")


async def check_summer_camp_payment_pending_2() -> None:
    """
    Scheduled job: 2nd follow-up for unpaid leads — 20 hours AFTER the first
    payment-pending message was sent (payment_followup_wa_sent_at).
    Sequential dependency: only fires if the 5-min message was already sent.
    Campaign: 'summercamp payment pending followup 1'. No media, no params.
    """
    from .notifications import send_whatsapp_notification

    twenty_hr_ago = datetime.now(timezone.utc).timestamp() - (20 * 60 * 60)
    max_age = datetime.now(timezone.utc).timestamp() - (30 * 24 * 60 * 60)  # 30-day cap

    try:
        cursor = db.summer_camp_bookings.find(
            {
                "crm_status": "lead",
                "payment_followup_wa_sent": True,          # ← must have received step 1
                "payment_followup2_wa_sent": {"$ne": True},
            },
            {"_id": 0, "id": 1, "parent_phone": 1, "child_name": 1,
             "created_at": 1, "payment_followup_wa_sent_at": 1}
        )
        for booking in await cursor.to_list(length=100):
            try:
                # Measure 20h from when step 1 was sent, not from updated_at
                ref_raw = booking.get("payment_followup_wa_sent_at") or booking.get("created_at", "")
                ref_dt = _parse_created_dt(ref_raw)
                if not ref_dt:
                    continue
                # Skip if step 1 was sent less than 20h ago
                if ref_dt.timestamp() > twenty_hr_ago:
                    continue
                # Skip ancient leads (older than 30 days) to avoid spam
                created_dt = _parse_created_dt(booking.get("created_at", ""))
                if created_dt and created_dt.timestamp() < max_age:
                    continue
            except Exception:
                continue

            phone = booking.get("parent_phone", "")
            if not phone:
                continue

            await db.summer_camp_bookings.update_one(
                {"id": booking["id"]},
                {"$set": {
                    "payment_followup2_wa_sent": True,
                    "payment_followup2_wa_sent_at": datetime.now(timezone.utc).isoformat(),
                }}
            )
            result = await send_whatsapp_notification(
                phone=phone,
                template_key="summercamp_payment_pending_2",
                params=[],
                user_name=booking.get("child_name", ""),
            )
            print(f"[SC PayPending2] {'Sent' if result.get('success') else 'Failed'} for {phone} (id={booking['id']})")

    except Exception as e:
        print(f"[SC PayPending2] Scheduler error: {e}")


async def check_summer_camp_payment_pending_3() -> None:
    """
    Scheduled job: 3rd follow-up for unpaid leads — 48 hours AFTER the 20-hour
    message was sent (payment_followup2_wa_sent_at).
    Sequential dependency: only fires if step 2 was already sent.
    Campaign: 'summer camp payment pending followup 2'. $FirstName param, no media.
    """
    from .notifications import send_whatsapp_notification

    forty_eight_hr_ago = datetime.now(timezone.utc).timestamp() - (48 * 60 * 60)
    max_age = datetime.now(timezone.utc).timestamp() - (30 * 24 * 60 * 60)

    try:
        cursor = db.summer_camp_bookings.find(
            {
                "crm_status": "lead",
                "payment_followup2_wa_sent": True,          # ← must have received step 2
                "payment_followup3_wa_sent": {"$ne": True},
            },
            {"_id": 0, "id": 1, "parent_phone": 1, "child_name": 1, "parent_name": 1,
             "created_at": 1, "payment_followup2_wa_sent_at": 1}
        )
        for booking in await cursor.to_list(length=100):
            try:
                # Measure 48h from when step 2 was sent
                ref_raw = booking.get("payment_followup2_wa_sent_at") or booking.get("created_at", "")
                ref_dt = _parse_created_dt(ref_raw)
                if not ref_dt or ref_dt.timestamp() > forty_eight_hr_ago:
                    continue
                created_dt = _parse_created_dt(booking.get("created_at", ""))
                if created_dt and created_dt.timestamp() < max_age:
                    continue
            except Exception:
                continue

            phone = booking.get("parent_phone", "")
            if not phone:
                continue

            raw_name = booking.get("child_name") or booking.get("parent_name") or "Student"
            first_name = raw_name.strip().split()[0] if raw_name.strip() else "Student"

            await db.summer_camp_bookings.update_one(
                {"id": booking["id"]},
                {"$set": {
                    "payment_followup3_wa_sent": True,
                    "payment_followup3_wa_sent_at": datetime.now(timezone.utc).isoformat(),
                }}
            )
            result = await send_whatsapp_notification(
                phone=phone,
                template_key="summercamp_payment_pending_3",
                params=["$FirstName"],
                user_name=first_name,
            )
            print(f"[SC PayPending3] {'Sent' if result.get('success') else 'Failed'} for {phone} (id={booking['id']}, name={first_name})")

    except Exception as e:
        print(f"[SC PayPending3] Scheduler error: {e}")


async def check_summer_camp_phone_captured_24h() -> None:
    """
    Scheduled job: 2nd follow-up for phone-captured leads — 24 hours AFTER the
    first 5-min follow-up was sent (follow_up_wa_sent_at).
    Sequential dependency: only fires if the 5-min message was already sent.
    Campaign: 'summer camp phone captured followup 1'
    """
    from .notifications import send_whatsapp_notification

    twenty_four_hr_ago = datetime.now(timezone.utc).timestamp() - (24 * 60 * 60)
    max_age = datetime.now(timezone.utc).timestamp() - (30 * 24 * 60 * 60)

    try:
        cursor = db.summer_camp_bookings.find(
            {
                "crm_status": "phone_captured",
                "follow_up_wa_sent": True,                          # ← must have received 5-min msg
                "phone_captured_followup1_wa_sent": {"$ne": True},
            },
            {"_id": 0, "id": 1, "parent_phone": 1, "child_name": 1,
             "created_at": 1, "follow_up_wa_sent_at": 1}
        )
        for booking in await cursor.to_list(length=100):
            try:
                # Measure 24h from when the 5-min message was sent
                ref_raw = booking.get("follow_up_wa_sent_at") or booking.get("created_at", "")
                ref_dt = _parse_created_dt(ref_raw)
                if not ref_dt or ref_dt.timestamp() > twenty_four_hr_ago:
                    continue
                created_dt = _parse_created_dt(booking.get("created_at", ""))
                if created_dt and created_dt.timestamp() < max_age:
                    continue
            except Exception:
                continue

            phone = booking.get("parent_phone", "")
            if not phone:
                continue

            await db.summer_camp_bookings.update_one(
                {"id": booking["id"]},
                {"$set": {
                    "phone_captured_followup1_wa_sent": True,
                    "phone_captured_followup1_wa_sent_at": datetime.now(timezone.utc).isoformat(),
                }}
            )
            result = await send_whatsapp_notification(
                phone=phone,
                template_key="summercamp_phone_captured_24h",
                params=[],
                user_name=booking.get("child_name", ""),
            )
            print(f"[SC PhoneCap24h] {'Sent' if result.get('success') else 'Failed'} for {phone} (id={booking['id']})")

    except Exception as e:
        print(f"[SC PhoneCap24h] Scheduler error: {e}")


async def check_summer_camp_closing_7days() -> None:
    """
    Scheduled job: Final 'registrations closing' follow-up for ALL unconverted leads
    (both phone_captured AND details-filled/lead) fired 7 days after creation.
    No media, no template params. Max age cap: 30 days.
    Campaign: 'summer camp registraitons closing followup'
    """
    from .notifications import send_whatsapp_notification

    seven_days_ago = datetime.now(timezone.utc).timestamp() - (7 * 24 * 60 * 60)
    max_age = datetime.now(timezone.utc).timestamp() - (30 * 24 * 60 * 60)

    try:
        cursor = db.summer_camp_bookings.find(
            {
                "crm_status": {"$in": ["phone_captured", "lead", "hot_lead"]},
                "closing_followup_wa_sent": {"$ne": True},
            },
            {"_id": 0, "id": 1, "parent_phone": 1, "child_name": 1, "created_at": 1}
        )
        for booking in await cursor.to_list(length=200):
            try:
                created_dt = _parse_created_dt(booking.get("created_at"))
                if not created_dt:
                    continue
                # Must be at least 7 days old
                if created_dt.timestamp() > seven_days_ago:
                    continue
                # Skip leads older than 30 days (irrelevant / already stale)
                if created_dt.timestamp() < max_age:
                    continue
            except Exception:
                continue

            phone = booking.get("parent_phone", "")
            if not phone:
                continue

            await db.summer_camp_bookings.update_one(
                {"id": booking["id"]},
                {"$set": {
                    "closing_followup_wa_sent": True,
                    "closing_followup_wa_sent_at": datetime.now(timezone.utc).isoformat(),
                }}
            )
            result = await send_whatsapp_notification(
                phone=phone,
                template_key="summercamp_closing_7days",
                params=[],
                user_name=booking.get("child_name", ""),
            )
            print(f"[SC Closing7d] {'Sent' if result.get('success') else 'Failed'} for {phone} (id={booking['id']})")

    except Exception as e:
        print(f"[SC Closing7d] Scheduler error: {e}")


# ─── Email Follow-Up Helpers ───────────────────────────────────────────────────

BOOKING_URL = os.getenv("FRONTEND_URL", "https://oll.co") + "/summer-camp/book"
_SC_EMAIL_FROM = "OLL Summer Camp <noreply@oll.co>"

def _sc_email_html(child_first: str, heading: str, body_html: str, cta_label: str = "Complete Registration") -> str:
    return f"""<!DOCTYPE html>
<html><head><meta charset="utf-8"><meta name="viewport" content="width=device-width,initial-scale=1">
<style>
  body{{margin:0;padding:0;background:#0B1120;font-family:'Segoe UI',Arial,sans-serif;color:#E2E8F0}}
  .wrap{{max-width:560px;margin:32px auto;background:#111827;border-radius:16px;overflow:hidden;border:1px solid #1e2d3d}}
  .header{{background:linear-gradient(135deg,#0B1E3D 0%,#1a3a6e 100%);padding:32px 32px 24px;text-align:center}}
  .logo{{font-size:22px;font-weight:800;letter-spacing:1px;color:#38BDF8;margin-bottom:4px}}
  .subtitle{{color:#64748B;font-size:12px;text-transform:uppercase;letter-spacing:2px}}
  .body{{padding:32px}}
  h1{{font-size:22px;font-weight:700;color:#F1F5F9;margin:0 0 16px}}
  p{{font-size:15px;line-height:1.7;color:#94A3B8;margin:0 0 16px}}
  .cta{{display:block;width:fit-content;margin:24px auto 0;padding:14px 36px;background:#38BDF8;color:#0B1120;border-radius:30px;font-weight:700;font-size:15px;text-decoration:none}}
  .footer{{padding:20px 32px;border-top:1px solid #1e2d3d;font-size:12px;color:#475569;text-align:center}}
  .badge{{display:inline-block;background:#1e3a5f;color:#38BDF8;border-radius:20px;padding:4px 14px;font-size:12px;font-weight:600;margin-bottom:20px}}
</style></head>
<body>
  <div class="wrap">
    <div class="header">
      <div class="logo">OLL</div>
      <div class="subtitle">Summer Camp 2026</div>
    </div>
    <div class="body">
      <span class="badge">Summer Camp 2026</span>
      <h1>Hi {child_first}, {heading}</h1>
      {body_html}
      <a href="{BOOKING_URL}" class="cta">{cta_label} &rarr;</a>
    </div>
    <div class="footer">
      One Lakh Learners &bull; Mumbai &bull; <a href="https://oll.co" style="color:#38BDF8">oll.co</a><br>
      You're receiving this because you expressed interest in OLL Summer Camp 2026.
    </div>
  </div>
</body></html>"""


async def _send_sc_followup_email(booking: dict, step_key: str, subject: str, heading: str, body_html: str, cta: str = "Complete Registration") -> bool:
    """Send a summer camp follow-up email. Returns True if sent successfully."""
    email = (booking.get("parent_email") or "").strip()
    if not email or "@" not in email:
        return False

    child = booking.get("child_name") or booking.get("parent_name") or "there"
    first = child.strip().split()[0] if child.strip() else "there"
    html = _sc_email_html(first, heading, body_html, cta)

    try:
        from .shared import get_resend_api_key, SENDER_EMAIL
        import resend as _resend

        key = await get_resend_api_key()
        if key:
            _resend.api_key = key
            params = {"from": _SC_EMAIL_FROM, "to": [email], "subject": subject, "html": html}
            await asyncio.to_thread(_resend.Emails.send, params)
            print(f"[SC Email {step_key}] Sent via Resend to {email}")
            return True

        print(f"[SC Email {step_key}] Resend not configured — skipping {email}")
        return False

    except Exception as exc:
        print(f"[SC Email {step_key}] Failed for {email}: {exc}")
        return False


async def _run_sc_email_followup(
    step_key: str,
    min_age_seconds: int,
    max_age_seconds: int,
    subject: str,
    heading: str,
    body_html: str,
    cta: str = "Complete Registration",
) -> None:
    """
    Generic scheduler function for summer camp email follow-ups.
    Finds unconverted leads with `parent_email` that haven't received this step,
    and sends the email once they've been in the DB long enough.
    """
    now = datetime.now(timezone.utc).timestamp()
    cutoff_old = now - min_age_seconds
    cutoff_max = now - max_age_seconds

    UNCONVERTED = ["phone_captured", "lead", "hot_lead"]
    sent_flag   = f"email_{step_key}_sent"

    try:
        cursor = db.summer_camp_bookings.find(
            {
                "crm_status": {"$in": UNCONVERTED},
                "parent_email": {"$exists": True, "$nin": ["", None]},
                sent_flag: {"$ne": True},
            },
            {"_id": 0, "id": 1, "parent_email": 1, "child_name": 1, "parent_name": 1, "created_at": 1}
        )
        bookings = await cursor.to_list(length=200)

        for booking in bookings:
            created_dt = _parse_created_dt(booking.get("created_at"))
            if not created_dt:
                continue
            ts = created_dt.timestamp()
            if ts > cutoff_old:
                continue   # too new — not time yet
            if ts < cutoff_max:
                continue   # too old — skip stale leads

            # Mark FIRST to prevent duplicate sends
            await db.summer_camp_bookings.update_one(
                {"id": booking["id"]},
                {"$set": {
                    sent_flag: True,
                    f"email_{step_key}_sent_at": datetime.now(timezone.utc).isoformat(),
                }}
            )

            ok = await _send_sc_followup_email(booking, step_key, subject, heading, body_html, cta)
            print(f"[SC Email {step_key}] {'✓' if ok else '✗'} {booking.get('parent_email')} (id={booking['id']})")

    except Exception as exc:
        print(f"[SC Email {step_key}] Scheduler error: {exc}")


# ── Scheduled email jobs (called by APScheduler) ──────────────────────────────

async def check_sc_email_1h() -> None:
    """1-hour follow-up email after summer camp form fill."""
    await _run_sc_email_followup(
        step_key="1h",
        min_age_seconds=60 * 60,          # 1 hour
        max_age_seconds=30 * 24 * 60 * 60,
        subject="Your OLL Summer Camp spot is still available!",
        heading="your Summer Camp spot is waiting",
        body_html="""
            <p>You recently filled out an interest form for <strong>OLL Summer Camp 2026</strong> — 
            that's exciting! We'd love to have your child join us.</p>
            <p>Seats are filling up fast. Tap below to complete your registration and lock in your spot.</p>
            <p><strong>What your child gets:</strong> live coding sessions, robotics, AI projects, 
            and a certificate — all in just 5 days.</p>
        """,
        cta="Book My Spot Now",
    )


async def check_sc_email_24h() -> None:
    """24-hour follow-up email for unconverted summer camp leads."""
    await _run_sc_email_followup(
        step_key="24h",
        min_age_seconds=24 * 60 * 60,
        max_age_seconds=30 * 24 * 60 * 60,
        subject="⏳ Limited seats left — OLL Summer Camp 2026",
        heading="a few seats are still open",
        body_html="""
            <p>It's been a day since you showed interest in <strong>OLL Summer Camp 2026</strong>. 
            We just wanted to let you know — seats are limited and going fast.</p>
            <p>Past campers built games, programmed robots, and went home with hands-on AI skills. 
            Your child could be next.</p>
            <p>Don't let the spot slip away!</p>
        """,
        cta="Reserve Your Seat",
    )


async def check_sc_email_2d() -> None:
    """2-day follow-up email for unconverted summer camp leads."""
    await _run_sc_email_followup(
        step_key="2d",
        min_age_seconds=2 * 24 * 60 * 60,
        max_age_seconds=30 * 24 * 60 * 60,
        subject="Your child could be coding this summer 🚀",
        heading="imagine what your child could build",
        body_html="""
            <p>In just 5 days at <strong>OLL Summer Camp 2026</strong>, kids age 4–16 learn to code, 
            build robots, and explore AI — no prior experience needed.</p>
            <ul style="color:#94A3B8;font-size:15px;line-height:2;padding-left:20px">
              <li>Ages 4–8: Block coding &amp; creative storytelling</li>
              <li>Ages 9–12: Python, games &amp; hardware projects</li>
              <li>Ages 13–16: AI, machine learning &amp; app development</li>
            </ul>
            <p>Spots are limited. Registration takes under 2 minutes.</p>
        """,
        cta="Start Registration",
    )


async def check_sc_email_5d() -> None:
    """5-day follow-up email for unconverted summer camp leads."""
    await _run_sc_email_followup(
        step_key="5d",
        min_age_seconds=5 * 24 * 60 * 60,
        max_age_seconds=30 * 24 * 60 * 60,
        subject="Last reminder: OLL Summer Camp 2026",
        heading="we'd still love to have you join us",
        body_html="""
            <p>It's been 5 days since you expressed interest in <strong>OLL Summer Camp 2026</strong>. 
            If you've been on the fence, now is the time to decide — registrations close soon.</p>
            <p>Thousands of kids have already signed up. Secure the last few remaining seats 
            for your child before they're gone.</p>
            <p>If you have any questions, just reply to this email — we're happy to help.</p>
        """,
        cta="Complete Registration Now",
    )
